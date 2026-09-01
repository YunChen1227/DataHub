#!/usr/bin/env python3
"""Batch test GRSB route against records from Excel test cases."""

from __future__ import annotations

import argparse
import glob
import hashlib
import io
import json
import msoffcrypto
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("RELAY_BASE_URL", "http://aiszcloud.cn:8080")
APP_KEY = os.environ.get("GRSB_APP_KEY", "bhiuvx5m4ug9")
APP_SECRET = os.environ.get("GRSB_APP_SECRET", "47c2080cb4f3024172be96bda597b34f")
CASES_DIR = Path(__file__).resolve().parent.parent / "test" / "cases"


@dataclass
class Record:
    source: str
    row: int
    name: str
    id_card: str


@dataclass
class Result:
    source: str
    row: int
    name: str
    id_card: str
    http_status: int
    error_code: str
    body_code: str
    latency_ms: float | None
    outcome: str
    raw: str


def sign_x1(body: dict[str, str], secret: str) -> str:
    items = sorted((k, v) for k, v in body.items() if v)
    payload = "".join(k + v for k, v in items) + secret
    return hashlib.md5(payload.encode("utf-8")).hexdigest()


def load_cs2608(path: Path) -> list[Record]:
    with path.open("rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password="2608")
        out = io.BytesIO()
        office.decrypt(out)
        out.seek(0)
        wb = load_workbook(out, read_only=True, data_only=True)
        sh = wb.active
        records: list[Record] = []
        for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
            if not row or len(row) < 2:
                continue
            name = str(row[0]).strip() if row[0] is not None else ""
            id_card = str(row[1]).strip() if row[1] is not None else ""
            if not name or not id_card:
                continue
            records.append(Record("cs2608.xlsx", i, name, id_card))
        return records


def load_gjj(path: Path) -> list[Record]:
    wb = load_workbook(path, read_only=True, data_only=True)
    records: list[Record] = []
    seen: set[tuple[str, str]] = set()
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
            if i == 1 and row and str(row[0]).strip() in ("证件号", "身份证号"):
                continue
            if not row or len(row) < 2:
                continue
            id_card = str(row[0]).strip() if row[0] is not None else ""
            name = str(row[1]).strip() if row[1] is not None else ""
            if not name or not id_card:
                continue
            key = (id_card, name)
            if key in seen:
                continue
            seen.add(key)
            records.append(Record(path.name, i, name, id_card))
    return records


def load_all_records() -> list[Record]:
    records: list[Record] = []
    cs_path = CASES_DIR / "cs2608.xlsx"
    if cs_path.exists():
        records.extend(load_cs2608(cs_path))
    for path in sorted(CASES_DIR.glob("gjj*.xlsx")):
        records.extend(load_gjj(path))
    return records


def classify(error_code: str, body_code: str) -> str:
    if error_code == "0" and body_code == "001":
        return "found"
    if error_code == "0" and body_code == "999":
        return "not_found"
    if error_code == "0":
        return "other_ok"
    return "error"


def query_one(rec: Record, session: requests.Session) -> Result:
    body = {"name": rec.name, "idCard": rec.id_card}
    payload = {
        "encryptionType": 1,
        "appKey": APP_KEY,
        "sign": sign_x1(body, APP_SECRET),
        "body": body,
    }
    started = time.time()
    try:
        resp = session.post(
            f"{BASE_URL}/v1/openapi/zlx/querySrmxGRSB",
            json=payload,
            timeout=30,
        )
        raw = resp.text
        http_status = resp.status_code
    except Exception as exc:  # noqa: BLE001
        return Result(
            source=rec.source,
            row=rec.row,
            name=rec.name,
            id_card=rec.id_card,
            http_status=0,
            error_code="NETWORK",
            body_code="",
            latency_ms=None,
            outcome="error",
            raw=str(exc),
        )

    latency_ms = (time.time() - started) * 1000
    error_code = ""
    body_code = ""
    try:
        data = resp.json()
        head = data.get("head") or {}
        body_obj = data.get("body") or {}
        error_code = str(head.get("errorCode", ""))
        body_code = str(body_obj.get("code", ""))
        if isinstance(head.get("time"), (int, float)):
            latency_ms = float(head["time"])
    except Exception:  # noqa: BLE001
        error_code = "PARSE"
    return Result(
        source=rec.source,
        row=rec.row,
        name=rec.name,
        id_card=rec.id_card,
        http_status=http_status,
        error_code=error_code,
        body_code=body_code,
        latency_ms=latency_ms,
        outcome=classify(error_code, body_code),
        raw=raw,
    )


def run_batch(records: list[Record], workers: int, out_path: Path) -> dict[str, Any]:
    results: list[Result] = []
    total = len(records)
    done = 0
    started = time.time()

    with requests.Session() as session, ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(query_one, rec, session): rec for rec in records}
        for fut in as_completed(futures):
            result = fut.result()
            results.append(result)
            done += 1
            if done % 25 == 0 or done == total:
                elapsed = time.time() - started
                print(f"progress {done}/{total} ({done/total:.1%}) elapsed={elapsed:.0f}s", flush=True)

    results.sort(key=lambda r: (r.source, r.row))
    summary = summarize(results)
    payload = {"summary": summary, "results": [asdict(r) for r in results]}
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def summarize(results: list[Result]) -> dict[str, Any]:
    total = len(results)
    by_outcome: dict[str, int] = {}
    by_error: dict[str, int] = {}
    by_body: dict[str, int] = {}
    for r in results:
        by_outcome[r.outcome] = by_outcome.get(r.outcome, 0) + 1
        if r.error_code:
            by_error[r.error_code] = by_error.get(r.error_code, 0) + 1
        if r.body_code:
            by_body[r.body_code] = by_body.get(r.body_code, 0) + 1

    found = by_outcome.get("found", 0)
    not_found = by_outcome.get("not_found", 0)
    errors = by_outcome.get("error", 0) + by_outcome.get("other_ok", 0)
    api_ok = found + not_found + by_outcome.get("other_ok", 0)

    by_source: dict[str, dict[str, int]] = {}
    for r in results:
        src = by_source.setdefault(r.source, {"total": 0, "found": 0, "not_found": 0, "error": 0})
        src["total"] += 1
        if r.outcome == "found":
            src["found"] += 1
        elif r.outcome == "not_found":
            src["not_found"] += 1
        else:
            src["error"] += 1

    latencies = [r.latency_ms for r in results if r.latency_ms is not None]
    return {
        "total": total,
        "found": found,
        "not_found": not_found,
        "error": errors,
        "api_ok": api_ok,
        "found_rate": round(found / total * 100, 2) if total else 0,
        "not_found_rate": round(not_found / total * 100, 2) if total else 0,
        "error_rate": round(errors / total * 100, 2) if total else 0,
        "api_success_rate": round(api_ok / total * 100, 2) if total else 0,
        "by_outcome": by_outcome,
        "by_error_code": by_error,
        "by_body_code": by_body,
        "by_source": by_source,
        "latency_ms_avg": round(sum(latencies) / len(latencies), 1) if latencies else None,
        "latency_ms_p95": round(sorted(latencies)[int(len(latencies) * 0.95) - 1], 1) if latencies else None,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workers", type=int, default=5)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "test_res" / "grsb_xlsx_batch.json",
    )
    args = parser.parse_args()

    records = load_all_records()
    print(f"loaded {len(records)} records from Excel files")
    print(f"target: {BASE_URL} appKey={APP_KEY}")

    health = requests.get(f"{BASE_URL}/healthz", timeout=10)
    if health.status_code != 200:
        print(f"healthz failed: {health.status_code}")
        return 1

    summary = run_batch(records, args.workers, args.out)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"saved: {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
