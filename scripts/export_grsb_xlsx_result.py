#!/usr/bin/env python3
"""Export grsb_xlsx_batch.json to Excel files split by original source documents."""

from __future__ import annotations

import glob
import io
import json
import msoffcrypto
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

HEADERS = ("姓名", "身份证号", "缴费时间", "参保状态", "缴费基数", "缴费单位", "个人身份")

CBJFZT_MAP = {
    "1": "正常参保",
    "2": "暂停参保",
}


def map_cbjfzt(value: object) -> str | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    if text in CBJFZT_MAP:
        return CBJFZT_MAP[text]
    return text


def parse_range(raw: str) -> dict | None:
    try:
        body = json.loads(raw).get("body") or {}
        range_s = (body.get("result") or {}).get("range") or ""
        if not str(range_s).strip():
            return None
        obj = json.loads(range_s)
        if isinstance(obj, list):
            return obj[0] if obj else None
        return obj
    except Exception:
        return None


def result_row(name: str, id_card: str, raw: str | None) -> tuple:
    if not raw:
        return (name, id_card, None, None, None, None, None)
    data = parse_range(raw)
    if not data:
        return (name, id_card, None, None, None, None, None)
    return (
        data.get("xm") or name,
        data.get("sfz") or id_card,
        data.get("jfsj") or None,
        map_cbjfzt(data.get("cbjfzt")),
        data.get("jfjs") or None,
        data.get("jfdw") or None,
        data.get("grsf") or None,
    )


def build_lookup(results: list[dict]) -> dict[tuple[str, str, str], str]:
    """Map (source, id_card, name) -> raw response."""
    lookup: dict[tuple[str, str, str], str] = {}
    for item in results:
        key = (item["source"], item["id_card"], item["name"])
        lookup[key] = item["raw"]
    return lookup


def write_sheet(ws, rows: list[tuple[str, str]], lookup: dict[tuple[str, str, str], str], source: str) -> dict:
    ws.append(HEADERS)
    with_data = 0
    missing = 0
    for name, id_card in rows:
        raw = lookup.get((source, id_card, name))
        if raw is None:
            raw = lookup.get((source, id_card, name.strip()))
        if raw is None:
            missing += 1
        row = result_row(name, id_card, raw)
        ws.append(row)
        if row[2] is not None:
            with_data += 1
    return {"total": len(rows), "with_data": with_data, "missing_lookup": missing}


def load_cs2608_rows(path: Path) -> list[tuple[str, str]]:
    with path.open("rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password="2608")
        out = io.BytesIO()
        office.decrypt(out)
        out.seek(0)
        wb = load_workbook(out, read_only=True, data_only=True)
        sh = wb.active
        rows: list[tuple[str, str]] = []
        for row in sh.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            name = str(row[0]).strip() if row[0] is not None else ""
            id_card = str(row[1]).strip() if row[1] is not None else ""
            if name and id_card:
                rows.append((name, id_card))
        return rows


def load_gjj_sheet_rows(path: Path, sheet_name: str) -> list[tuple[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sh = wb[sheet_name]
    rows: list[tuple[str, str]] = []
    for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        if i == 1 and str(row[0]).strip() in ("证件号", "身份证号"):
            continue
        id_card = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if id_card and name:
            rows.append((name, id_card))
    return rows


def inspect_sources(cases_dir: Path) -> dict:
    report: dict = {"cs2608.xlsx": {}, "gjj": {}}

    cs_path = cases_dir / "cs2608.xlsx"
    if cs_path.exists():
        with cs_path.open("rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password="2608")
            out = io.BytesIO()
            office.decrypt(out)
            out.seek(0)
            wb = load_workbook(out, read_only=True, data_only=True)
            report["cs2608.xlsx"] = {
                "sheets": wb.sheetnames,
                "rows_per_sheet": {
                    name: len(load_cs2608_rows(cs_path)) if name == wb.sheetnames[0] else 0
                    for name in wb.sheetnames
                },
            }
        report["cs2608.xlsx"]["rows_per_sheet"] = {
            wb.sheetnames[0]: len(load_cs2608_rows(cs_path))
        }

    gjj_paths = sorted(cases_dir.glob("gjj*.xlsx"))
    if gjj_paths:
        gjj_path = gjj_paths[0]
        wb = load_workbook(gjj_path, read_only=True, data_only=True)
        report["gjj"] = {
            "file": gjj_path.name,
            "sheets": wb.sheetnames,
            "rows_per_sheet": {
                name: len(load_gjj_sheet_rows(gjj_path, name)) for name in wb.sheetnames
            },
        }
    return report


def export_split(json_path: Path, out_dir: Path) -> dict:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    lookup = build_lookup(payload["results"])
    cases_dir = json_path.parent.parent / "test" / "cases"
    out_dir.mkdir(parents=True, exist_ok=True)

    source_report = inspect_sources(cases_dir)
    summary: dict = {"source_inspection": source_report, "files": {}}

    # cs2608
    cs_path = cases_dir / "cs2608.xlsx"
    cs_rows = load_cs2608_rows(cs_path)
    wb_cs = Workbook()
    ws_cs = wb_cs.active
    ws_cs.title = "BJPG-01"
    cs_out = out_dir / "cs2608_社保查询结果.xlsx"
    summary["files"]["cs2608_社保查询结果.xlsx"] = write_sheet(
        ws_cs, cs_rows, lookup, "cs2608.xlsx"
    )
    wb_cs.save(cs_out)

    # gjj (all sheets)
    gjj_path = sorted(cases_dir.glob("gjj*.xlsx"))[0]
    gjj_out = out_dir / "gjj测试0826_社保查询结果.xlsx"
    wb_gjj = Workbook()
    wb_gjj.remove(wb_gjj.active)
    gjj_source = gjj_path.name
    summary["files"]["gjj测试0826_社保查询结果.xlsx"] = {"sheets": {}}
    for sheet_name in load_workbook(gjj_path, read_only=True).sheetnames:
        rows = load_gjj_sheet_rows(gjj_path, sheet_name)
        ws = wb_gjj.create_sheet(sheet_name)
        sheet_summary = write_sheet(ws, rows, lookup, gjj_source)
        summary["files"]["gjj测试0826_社保查询结果.xlsx"]["sheets"][sheet_name] = sheet_summary
    wb_gjj.save(gjj_out)

    summary["output_dir"] = str(out_dir)
    return summary


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    json_path = root / "test_res" / "grsb_xlsx_batch.json"
    out_dir = root / "test_res"

    if not json_path.exists():
        print(f"missing input: {json_path}", file=sys.stderr)
        return 1

    summary = export_split(json_path, out_dir)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
