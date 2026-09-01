#!/usr/bin/env python3
"""Backfill missing GRSB tests and export two complete result workbooks."""

from __future__ import annotations

import glob
import hashlib
import io
import json
import msoffcrypto
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("RELAY_BASE_URL", "http://aiszcloud.cn:8080")
APP_KEY = os.environ.get("GRSB_APP_KEY", "bhiuvx5m4ug9")
APP_SECRET = os.environ.get("GRSB_APP_SECRET", "47c2080cb4f3024172be96bda597b34f")

HEADERS = ("姓名", "身份证号", "缴费时间", "参保状态", "缴费基数", "缴费单位", "个人身份")
CBJFZT_MAP = {"1": "正常参保", "2": "暂停参保"}

ROOT = Path(__file__).resolve().parent.parent
CASES_DIR = ROOT / "test" / "cases"
JSON_PATH = ROOT / "test_res" / "grsb_xlsx_batch.json"
OUT_CS = ROOT / "test_res" / "cs2608_社保查询结果_完整.xlsx"
OUT_GJJ = ROOT / "test_res" / "gjj测试0826_社保查询结果_完整.xlsx"


@dataclass
class SourceRow:
    source: str
    sheet: str | None
    row_num: int
    name: str
    id_card: str


@dataclass
class QueryResult:
    source: str
    sheet: str | None
    row_num: int
    name: str
    id_card: str
    http_status: int
    error_code: str
    body_code: str
    latency_ms: float | None
    outcome: str
    raw: str


def sign_x1(body: dict[str, str], secret: str) -> str:
    items = sorted((k, v) for k, v in body.items() if v)
    payload = "".join(k + v for k, v in items) + secret
    return hashlib.md5(payload.encode("utf-8")).hexdigest()


def map_cbjfzt(value: object) -> str | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    return CBJFZT_MAP.get(text, text)


def parse_range(raw: str) -> dict | None:
    try:
        body = json.loads(raw).get("body") or {}
        range_s = (body.get("result") or {}).get("range") or ""
        if not str(range_s).strip():
            return None
        obj = json.loads(range_s)
        if isinstance(obj, list):
            return obj[0] if obj else None
        return obj
    except Exception:
        return None


def classify(error_code: str, body_code: str) -> str:
    if error_code == "0" and body_code == "001":
        return "found"
    if error_code == "0" and body_code == "999":
        return "not_found"
    if error_code == "0":
        return "other_ok"
    return "error"


def load_cs2608_rows() -> list[SourceRow]:
    path = CASES_DIR / "cs2608.xlsx"
    with path.open("rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password="2608")
        out = io.BytesIO()
        office.decrypt(out)
        out.seek(0)
        sh = load_workbook(out, read_only=True, data_only=True).active
        rows: list[SourceRow] = []
        for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
            if not row or len(row) < 2:
                continue
            name = str(row[0]).strip() if row[0] is not None else ""
            id_card = str(row[1]).strip() if row[1] is not None else ""
            if not id_card:
                continue
            rows.append(SourceRow("cs2608.xlsx", None, i, name, id_card))
        return rows


def load_gjj_rows() -> dict[str, list[SourceRow]]:
    path = Path(glob.glob(str(CASES_DIR / "gjj*.xlsx"))[0])
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[SourceRow]] = {}
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        rows: list[SourceRow] = []
        for i, row in enumerate(sh.iter_rows(values_only=True), start=1):
            if not any(c is not None and str(c).strip() for c in row):
                continue
            if i == 1 and str(row[0]).strip() in ("证件号", "身份证号"):
                continue
            id_card = str(row[0]).strip() if row[0] is not None else ""
            name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not id_card:
                continue
            rows.append(SourceRow(path.name, sheet_name, i, name, id_card))
        out[sheet_name] = rows
    return out


def result_key(source: str, sheet: str | None, id_card: str, name: str) -> tuple:
    return (source, sheet or "", id_card, name)


def load_existing_results() -> dict[tuple, QueryResult]:
    payload = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    out: dict[tuple, QueryResult] = {}
    for item in payload["results"]:
        sheet = item.get("sheet")
        key = result_key(item["source"], sheet, item["id_card"], item["name"])
        out[key] = QueryResult(
            source=item["source"],
            sheet=sheet,
            row_num=item.get("row", 0),
            name=item["name"],
            id_card=item["id_card"],
            http_status=item["http_status"],
            error_code=item["error_code"],
            body_code=item["body_code"],
            latency_ms=item.get("latency_ms"),
            outcome=item["outcome"],
            raw=item["raw"],
        )
    return out


def lookup_result(
    lookup: dict[tuple, QueryResult],
    by_id: dict[str, QueryResult],
    row: SourceRow,
) -> QueryResult | None:
    keys = [
        result_key(row.source, row.sheet, row.id_card, row.name),
        result_key(row.source, None, row.id_card, row.name),
        result_key(row.source, row.sheet, row.id_card, ""),
    ]
    for key in keys:
        if key in lookup:
            return lookup[key]
    if row.id_card in by_id:
        return by_id[row.id_card]
    return None


def query_one(row: SourceRow, session: requests.Session) -> QueryResult:
    body = {"name": row.name, "idCard": row.id_card}
    payload = {
        "encryptionType": 1,
        "appKey": APP_KEY,
        "sign": sign_x1(body, APP_SECRET),
        "body": body,
    }
    started = time.time()
    try:
        resp = session.post(
            f"{BASE_URL}/v1/openapi/zlx/querySrmxGRSB",
            json=payload,
            timeout=30,
        )
        raw = resp.text
        http_status = resp.status_code
    except Exception as exc:  # noqa: BLE001
        return QueryResult(
            row.source, row.sheet, row.row_num, row.name, row.id_card,
            0, "NETWORK", "", None, "error", str(exc),
        )

    latency_ms = (time.time() - started) * 1000
    error_code = ""
    body_code = ""
    try:
        data = resp.json()
        head = data.get("head") or {}
        body_obj = data.get("body") or {}
        error_code = str(head.get("errorCode", ""))
        body_code = str(body_obj.get("code", ""))
        if isinstance(head.get("time"), (int, float)):
            latency_ms = float(head["time"])
    except Exception:  # noqa: BLE001
        error_code = "PARSE"
    return QueryResult(
        row.source, row.sheet, row.row_num, row.name, row.id_card,
        http_status, error_code, body_code, latency_ms,
        classify(error_code, body_code), raw,
    )


def backfill_missing(rows: list[SourceRow], lookup: dict[tuple, QueryResult], by_id: dict[str, QueryResult]) -> list[QueryResult]:
    missing: list[SourceRow] = []
    for row in rows:
        if lookup_result(lookup, by_id, row) is None:
            missing.append(row)
    if not missing:
        return []

    print(f"backfill {len(missing)} missing rows...", flush=True)
    new_results: list[QueryResult] = []
    with requests.Session() as session, ThreadPoolExecutor(max_workers=5) as pool:
        futures = {pool.submit(query_one, row, session): row for row in missing}
        for fut in as_completed(futures):
            result = fut.result()
            new_results.append(result)
            key = result_key(result.source, result.sheet, result.id_card, result.name)
            lookup[key] = result
            by_id[result.id_card] = result
    return new_results


def result_row_from_query(row: SourceRow, result: QueryResult | None) -> tuple:
    display_name = row.name or None
    if result is None:
        return (display_name, row.id_card, None, None, None, None, None)
    data = parse_range(result.raw)
    if not data:
        return (display_name, row.id_card, None, None, None, None, None)
    return (
        data.get("xm") or display_name,
        data.get("sfz") or row.id_card,
        data.get("jfsj") or None,
        map_cbjfzt(data.get("cbjfzt")),
        data.get("jfjs") or None,
        data.get("jfdw") or None,
        data.get("grsf") or None,
    )


def write_workbook(path: Path, sheets: dict[str, list[tuple]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        for row in rows:
            ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def save_merged_json(lookup: dict[tuple, QueryResult]) -> None:
    results = sorted(
        lookup.values(),
        key=lambda r: (r.source, r.sheet or "", r.row_num, r.id_card),
    )
    payload = {
        "summary": {"total": len(results)},
        "results": [asdict(r) for r in results],
    }
    JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    cs_rows = load_cs2608_rows()
    gjj_rows = load_gjj_rows()
    lookup = load_existing_results()
    by_id = {r.id_card: r for r in lookup.values()}

    # annotate sheet on legacy gjj rows without sheet
    for r in list(lookup.values()):
        if "gjj" in r.source and not r.sheet:
            r.sheet = "浙江"

    all_rows = cs_rows + [r for sheet in gjj_rows for r in gjj_rows[sheet]]
    new_results = backfill_missing(all_rows, lookup, by_id)
    save_merged_json(lookup)

    # Build output rows in exact original order
    cs_out: list[tuple] = []
    cs_with_data = 0
    for row in cs_rows:
        out_row = result_row_from_query(row, lookup_result(lookup, by_id, row))
        cs_out.append(out_row)
        if out_row[2] is not None:
            cs_with_data += 1

    gjj_out: dict[str, list[tuple]] = {}
    gjj_stats: dict[str, dict] = {}
    gjj_path_name = Path(glob.glob(str(CASES_DIR / "gjj*.xlsx"))[0]).name
    for sheet_name, rows in gjj_rows.items():
        sheet_rows: list[tuple] = []
        with_data = 0
        for row in rows:
            row.source = gjj_path_name
            out_row = result_row_from_query(row, lookup_result(lookup, by_id, row))
            sheet_rows.append(out_row)
            if out_row[2] is not None:
                with_data += 1
        gjj_out[sheet_name] = sheet_rows
        gjj_stats[sheet_name] = {"total": len(sheet_rows), "with_data": with_data}

    write_workbook(OUT_CS, {"BJPG-01": cs_out})
    write_workbook(OUT_GJJ, gjj_out)

    summary = {
        "backfilled": len(new_results),
        "cs2608": {"total": len(cs_out), "with_data": cs_with_data, "output": str(OUT_CS)},
        "gjj": {"sheets": gjj_stats, "output": str(OUT_GJJ)},
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
